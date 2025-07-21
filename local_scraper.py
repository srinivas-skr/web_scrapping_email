# ==============================================================================
# FINAL, CORRECTED & EFFICIENT SCRIPT (Production-Grade v5.0)
# This version fixes the critical performance flaw while maintaining incremental saving.
# ==============================================================================

import os
import time
import re
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl  # Directly using openpyxl for efficient incremental saves
from openpyxl.utils import get_column_letter

# Disable SSL warnings
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# --- Configuration ---
URLS_FILE         = 'urls.txt'
OUTPUT_XLSX       = 'extracted_emails.xlsx'
EMAIL_REGEX       = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
BLACKLIST_DOMAINS = {
    'sentry.wixpress.com',
    'cdn.ampproject.org',
}

# --- Helper Functions ---

def normalize_url(url: str) -> str:
    """Ensures a URL has a proper http/https scheme."""
    if not url.startswith(('http://', 'https://')):
        print(f"  > Info: Normalizing URL: http://{url}")
        return 'http://' + url
    return url

def find_emails_in_text(text: str, source_url: str) -> dict:
    emails = {}
    for raw in re.findall(EMAIL_REGEX, text):
        email = raw.lower()
        local, domain = email.split('@', 1)

        if email.endswith(('.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.woff')):
            continue
        if len(local) >= 16 and re.fullmatch(r'[0-9a-f]+', local):
            continue
        if domain in BLACKLIST_DOMAINS:
            continue

        score = 0
        if any(k in source_url.lower() for k in ('contact', 'about', 'connect')):
            score += 3
        if local in ('info', 'support', 'sales', 'hello', 'admin', 'contact',
                     'enquiries', 'office', 'service', 'mail', 'team', 'help'):
            score -= 1
        base_dom = urlparse(source_url).netloc.replace('www.', '')
        if base_dom and base_dom in email:
            score += 5
        if any(w in local for w in ('official', 'company', 'corporate', 'business')):
            score += 2

        emails[email] = max(emails.get(email, 0), score)
    return emails

def extract_additional_emails(soup: BeautifulSoup, source_url: str) -> dict:
    emails = {}
    for a in soup.find_all('a', href=True):
        if a['href'].lower().startswith('mailto:'):
            addr = a['href'].split(':', 1)[1].split('?')[0]
            emails.update(find_emails_in_text(addr, source_url))
    for img in soup.find_all('img', alt=True):
        emails.update(find_emails_in_text(img['alt'], source_url))
    return emails

def get_relevant_links(soup: BeautifulSoup, base_url: str) -> list:
    keywords = ('contact', 'about', 'team', 'support', 'enquiry', 'help', 'info')
    links = set()
    for a in soup.find_all('a', href=True):
        low = a['href'].lower()
        if any(k in low for k in keywords):
            full = urljoin(base_url, a['href'])
            if urlparse(full).netloc == urlparse(base_url).netloc:
                links.add(full)
    return list(links)

# --- Scraping Methods ---

def try_fast_method(url: str) -> dict:
    emails = {}
    try:
        sess = requests.Session()
        sess.headers.update({'User-Agent': 'Mozilla/5.0'})
        resp = sess.get(url, timeout=10, verify=False)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')

        emails.update(find_emails_in_text(resp.text, url))
        emails.update(extract_additional_emails(soup, url))

        for link in get_relevant_links(soup, url):
            try:
                sub = sess.get(link, timeout=8, verify=False)
                sub.raise_for_status()
                sub_soup = BeautifulSoup(sub.text, 'html.parser')
                emails.update(find_emails_in_text(sub.text, link))
                emails.update(extract_additional_emails(sub_soup, link))
            except Exception:
                pass
    except Exception:
        print(f"  > Fast method failed for {url}. Trying Selenium.")
    return emails

def try_selenium_method(url: str) -> dict:
    emails = {}
    driver = None
    try:
        opts = Options()
        opts.add_argument("--headless")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--log-level=3")
        opts.add_argument("user-agent=Mozilla/5.0")
        opts.add_experimental_option('excludeSwitches', ['enable-logging'])

        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=opts
        )
        driver.get(url)
        time.sleep(3)
        src = driver.page_source
        soup = BeautifulSoup(src, 'html.parser')

        emails.update(find_emails_in_text(src, url))
        emails.update(extract_additional_emails(soup, url))

        for link in get_relevant_links(soup, url):
            try:
                driver.get(link)
                time.sleep(2)
                page2 = driver.page_source
                soup2 = BeautifulSoup(page2, 'html.parser')
                emails.update(find_emails_in_text(page2, link))
                emails.update(extract_additional_emails(soup2, link))
            except Exception:
                pass
    except Exception as e:
        print(f"  > Selenium method failed for {url}: {e}")
    finally:
        if driver:
            driver.quit()
    return emails

# --- Main Logic with EFFICIENT Incremental Saving ---

def main():
    if not os.path.exists(URLS_FILE):
        print(f"⚠️ '{URLS_FILE}' not found. Create it with one URL per line.")
        return

    with open(URLS_FILE) as f:
        urls = [normalize_url(u.strip()) for u in f if u.strip()]

    print(f"✅ Loaded {len(urls)} URLs.")
    total_found = 0

    # Create the Excel file and header if it doesn't exist
    if not os.path.exists(OUTPUT_XLSX):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Email_Results"
        ws.append(['#', 'URL', 'Emails'])
        wb.save(OUTPUT_XLSX)

    for idx, url in enumerate(urls, start=1):
        print(f"\n[{idx}/{len(urls)}] Scraping: {url}")
        found = try_fast_method(url) or try_selenium_method(url)

        if found:
            sorted_em = sorted(found.items(), key=lambda x: x[1], reverse=True)
            unique = [email for email, _ in sorted_em]
            top = unique[:2]
            if top:
                emails_str = ', '.join(top)
                print(f"  ✅ Found: {emails_str}")
                total_found += len(top)
            else:
                emails_str = 'No valid company emails found'
                print("  ⚠️ No valid company emails after filtering")
        else:
            emails_str = 'No email found'
            print("  ⚠️ No emails found")

        # EFFICIENT INCREMENTAL SAVE
        try:
            wb = openpyxl.load_workbook(OUTPUT_XLSX)
            ws = wb.active
            ws.append([idx, url, emails_str])
            wb.save(OUTPUT_XLSX)
        except Exception as e:
            print(f"  ❌ ERROR saving to Excel. Is the file open? Error: {e}")

        time.sleep(1)

    # Final column auto‑size
    print(f"\n🏁 Done! Polishing final Excel file...")
    wb = openpyxl.load_workbook(OUTPUT_XLSX)
    ws = wb.active
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    wb.save(OUTPUT_XLSX)

    print(f"✅ {total_found} emails found. Results saved continuously to → {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
